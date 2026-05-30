import csv
import logging
import re
import sqlite3
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from xml.sax.saxutils import escape


DEFAULT_CATEGORY_RULES = {
    "all inkl": "Webhosting",
    "all-inkl": "Webhosting",
    "apotheke": "Gesundheit",
    "aral": "KFZ",
    "arzt": "Gesundheit",
    "artz": "Gesundheit",
    "auto": "KFZ",
    "autoversicherung": "Versicherung",
    "avia": "KFZ",
    "congstar": "Telekommunikation",
    "devk": "Versicherung",
    "duinrell": "Freizeit",
    "elster": "Steuer",
    "esso": "KFZ",
    "eurowings": "Reisekosten",
    "finanzamt": "Steuer",
    "gebäudeversicherung": "Versicherung",
    "gebaeudeversicherung": "Versicherung",
    "grundbesitz": "Immobilie",
    "hausrat": "Versicherung",
    "hem": "KFZ",
    "hotel": "Reisekosten",
    "huk24": "Versicherung",
    "jet": "KFZ",
    "kfz": "KFZ",
    "kraftstoff": "KFZ",
    "lkh": "Versicherung",
    "porsche": "KFZ",
    "porsche zentrum": "KFZ",
    "restaurant": "Bewirtung",
    "rechtschutz": "Versicherung",
    "rechtsschutz": "Versicherung",
    "sc technology": "Beratung/Abrechnung",
    "shell": "KFZ",
    "star": "KFZ",
    "strato": "Webhosting",
    "strom": "Versorgung",
    "tankbeleg": "KFZ",
    "tankbelege": "KFZ",
    "tankstelle": "KFZ",
    "totalenergies": "KFZ",
    "vodafone": "Telekommunikation",
    "wasser": "Versorgung",
    "zahn": "Gesundheit",
    "zahnarzt": "Gesundheit",
}

REVIEW_KEYWORDS = (
    "angebot",
    "anordnung",
    "betriebsprüfung",
    "bescheinigung",
    "elster",
    "login",
    "meldebescheinigung",
    "lohnsteuerbescheinigung",
    "uvg",
)

DETAIL_HEADERS = (
    "Lfd. Nr.",
    "Datum",
    "Anbieter",
    "Betrag",
    "Waehrung",
    "Steuerkategorie",
    "Hinweis",
    "Status",
    "Rechnungsnummer",
    "Archivpfad",
    "Quelle",
)

SUMMARY_HEADERS = ("Steuerkategorie", "Anzahl", "Summe EUR", "Anteil %")
REVIEW_HEADERS = DETAIL_HEADERS
VENDOR_HEADERS = ("Anbieter", "Steuerkategorie (haeufigste)", "Anzahl", "Summe EUR")
QUARTER_HEADERS = ("Quartal", "Steuerkategorie", "Anzahl", "Summe EUR")


@dataclass
class TaxExportConfig:
    database_path: Path
    output_dir: Path
    category_rules: dict[str, str]


@dataclass
class TaxExportResult:
    year: int
    rows: int
    review_rows: int
    csv_path: Path
    xlsx_path: Path


def export_tax_year(config: TaxExportConfig, year: int) -> TaxExportResult:
    rows = _load_invoice_rows(config.database_path, year)
    categorized = [_categorize(row, config.category_rules) for row in rows]

    year_dir = config.output_dir / str(year)
    year_dir.mkdir(parents=True, exist_ok=True)
    csv_path = year_dir / f"einkommensteuer_{year}.csv"
    xlsx_path = year_dir / f"einkommensteuer_{year}.xlsx"

    detail_rows = [DETAIL_HEADERS] + [_detail_row(i, row) for i, row in enumerate(categorized, start=1)]
    summary_rows = [SUMMARY_HEADERS] + _summary_rows(categorized)
    review_rows = [REVIEW_HEADERS] + [
        _detail_row(i, row)
        for i, row in enumerate(
            (r for r in categorized if r["tax_category"] == "Review"), start=1
        )
    ]
    monthly_rows = _monthly_rows(categorized)
    quarterly_rows = _quarterly_rows(categorized)
    vendor_rows = _vendor_rows(categorized)
    overview_rows = _overview_rows(year, categorized)

    _write_csv(csv_path, detail_rows)

    sheets = {
        "Uebersicht": overview_rows,
        "Alle Belege": detail_rows,
        "Monate": monthly_rows,
        "Quartale": quarterly_rows,
        "Kategorien": summary_rows,
        "Anbieter": vendor_rows,
        "Review": review_rows,
    }

    if not _write_xlsx_styled(xlsx_path, sheets, year):
        _write_xlsx(xlsx_path, sheets)

    return TaxExportResult(
        year=year,
        rows=len(categorized),
        review_rows=len(review_rows) - 1,
        csv_path=csv_path,
        xlsx_path=xlsx_path,
    )


def _load_invoice_rows(database_path: Path, year: int) -> list[sqlite3.Row]:
    con = sqlite3.connect(database_path)
    con.row_factory = sqlite3.Row
    try:
        return con.execute(
            """
            select invoice_date, vendor, amount, currency, invoice_number, status,
                   source_path, archive_path, category, reason
            from invoices
            where substr(invoice_date, 1, 4) = ?
            order by invoice_date, vendor
            """,
            (str(year),),
        ).fetchall()
    finally:
        con.close()


def _categorize(row: sqlite3.Row, rules: dict[str, str]) -> dict:
    text = " ".join(
        str(row[key] or "")
        for key in ("vendor", "source_path", "archive_path", "category", "reason")
    ).lower()

    tax_category = ""
    notes: list[str] = []

    # Spezifischstes (längstes) Keyword gewinnt, nicht das erste in Dict-Reihenfolge.
    matched = [
        (keyword, category)
        for keyword, category in rules.items()
        if re.search(rf"(?<![a-z0-9]){re.escape(keyword.lower())}(?![a-z0-9])", text)
    ]
    if matched:
        keyword, category = max(matched, key=lambda kc: len(kc[0]))
        tax_category = category
        notes.append(f"Regel: {keyword}")
    else:
        tax_category = "Review"
        notes.append("Keine Regel gefunden")

    if row["status"] != "archived":
        tax_category = "Review"
        notes.append(f"Status: {row['status']}")

    if row["amount"] is None:
        tax_category = "Review"
        notes.append("Betrag fehlt")

    review_matches = [kw for kw in REVIEW_KEYWORDS if kw in text]
    if review_matches:
        tax_category = "Review"
        notes.append("Review-Schluesselwort: " + ", ".join(review_matches[:3]))

    return {
        "invoice_date": row["invoice_date"],
        "vendor": row["vendor"],
        "amount": row["amount"],
        "currency": row["currency"],
        "invoice_number": row["invoice_number"] or "",
        "status": row["status"],
        "source_path": row["source_path"],
        "archive_path": row["archive_path"],
        "tax_category": tax_category,
        "note": "; ".join(notes),
    }


def _detail_row(seq: int, row: dict) -> list[object]:
    invoice_date = _to_date(row["invoice_date"])
    amount = row["amount"] if row["amount"] is not None else ""
    return [
        seq,
        invoice_date,
        row["vendor"],
        amount,
        row["currency"],
        row["tax_category"],
        row["note"],
        row["status"],
        row["invoice_number"],
        row["archive_path"],
        row["source_path"],
    ]


def _to_date(value):
    if isinstance(value, date):
        return value
    if not value:
        return ""
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return str(value)


def _summary_rows(rows: Iterable[dict]) -> list[list[object]]:
    grouped: dict[str, dict[str, float]] = {}
    total_sum = 0.0
    for row in rows:
        if row["currency"] != "EUR" or row["amount"] is None:
            continue
        bucket = grouped.setdefault(row["tax_category"], {"count": 0, "sum": 0.0})
        bucket["count"] += 1
        bucket["sum"] += float(row["amount"])
        total_sum += float(row["amount"])

    output = []
    # Sortierung: groesste Summe zuerst (interessant fuer den Steuerberater)
    for category in sorted(grouped, key=lambda k: grouped[k]["sum"], reverse=True):
        bucket = grouped[category]
        share = round(bucket["sum"] / total_sum * 100, 2) if total_sum > 0 else 0.0
        output.append([category, bucket["count"], round(bucket["sum"], 2), share])
    return output


MONTHLY_HEADERS = ("Monat", "Steuerkategorie", "Anzahl", "Summe EUR")


def _monthly_rows(rows: Iterable[dict]) -> list[list[object]]:
    grouped: dict[tuple[str, str], dict[str, float]] = {}
    for row in rows:
        if row["currency"] != "EUR" or row["amount"] is None:
            continue
        d = _to_date(row["invoice_date"])
        if not isinstance(d, date):
            continue
        month = f"{d.year:04d}-{d.month:02d}"
        key = (month, row["tax_category"])
        bucket = grouped.setdefault(key, {"count": 0, "sum": 0.0})
        bucket["count"] += 1
        bucket["sum"] += float(row["amount"])

    output: list[list[object]] = [list(MONTHLY_HEADERS)]
    for (month, category) in sorted(grouped):
        bucket = grouped[(month, category)]
        output.append([month, category, bucket["count"], round(bucket["sum"], 2)])
    return output


def _quarterly_rows(rows: Iterable[dict]) -> list[list[object]]:
    grouped: dict[tuple[str, str], dict[str, float]] = {}
    for row in rows:
        if row["currency"] != "EUR" or row["amount"] is None:
            continue
        d = _to_date(row["invoice_date"])
        if not isinstance(d, date):
            continue
        quarter = f"{d.year:04d}-Q{(d.month - 1) // 3 + 1}"
        key = (quarter, row["tax_category"])
        bucket = grouped.setdefault(key, {"count": 0, "sum": 0.0})
        bucket["count"] += 1
        bucket["sum"] += float(row["amount"])

    output: list[list[object]] = [list(QUARTER_HEADERS)]
    for (quarter, category) in sorted(grouped):
        bucket = grouped[(quarter, category)]
        output.append([quarter, category, bucket["count"], round(bucket["sum"], 2)])
    return output


def _vendor_rows(rows: Iterable[dict]) -> list[list[object]]:
    grouped: dict[str, dict] = {}
    for row in rows:
        if row["currency"] != "EUR" or row["amount"] is None:
            continue
        vendor = row["vendor"] or "(unbekannt)"
        bucket = grouped.setdefault(
            vendor, {"count": 0, "sum": 0.0, "categories": Counter()}
        )
        bucket["count"] += 1
        bucket["sum"] += float(row["amount"])
        bucket["categories"][row["tax_category"]] += 1

    output: list[list[object]] = [list(VENDOR_HEADERS)]
    for vendor in sorted(grouped, key=lambda k: grouped[k]["sum"], reverse=True):
        bucket = grouped[vendor]
        top_category = bucket["categories"].most_common(1)[0][0]
        output.append([vendor, top_category, bucket["count"], round(bucket["sum"], 2)])
    return output


OVERVIEW_HEADERS = ("Kennzahl", "Wert")


def _overview_rows(year: int, rows: list[dict]) -> list[list[object]]:
    total_count = len(rows)
    archived = sum(1 for r in rows if r["status"] == "archived")
    review = sum(1 for r in rows if r["tax_category"] == "Review")
    eur_amounts = [
        float(r["amount"]) for r in rows if r["currency"] == "EUR" and r["amount"] is not None
    ]
    eur_amounts_clean = [
        float(r["amount"])
        for r in rows
        if r["currency"] == "EUR"
        and r["amount"] is not None
        and r["tax_category"] != "Review"
    ]
    eur_sum = round(sum(eur_amounts), 2)
    eur_sum_clean = round(sum(eur_amounts_clean), 2)
    missing_amount = sum(1 for r in rows if r["amount"] is None)
    highest = max(eur_amounts) if eur_amounts else 0.0
    average = round(sum(eur_amounts) / len(eur_amounts), 2) if eur_amounts else 0.0

    cat_sum: Counter = Counter()
    vendor_sum: Counter = Counter()
    for r in rows:
        if r["currency"] != "EUR" or r["amount"] is None:
            continue
        cat_sum[r["tax_category"]] += float(r["amount"])
        vendor_sum[r["vendor"] or "(unbekannt)"] += float(r["amount"])
    top_category = cat_sum.most_common(1)[0][0] if cat_sum else "-"
    top_vendor = vendor_sum.most_common(1)[0][0] if vendor_sum else "-"

    return [
        list(OVERVIEW_HEADERS),
        ["Steuerjahr", year],
        ["Belege gesamt", total_count],
        ["Davon archiviert", archived],
        ["Davon zur Pruefung (Review)", review],
        ["Belege ohne Betrag", missing_amount],
        ["Summe EUR (alle Belege)", eur_sum],
        ["Summe EUR (ohne Review)", eur_sum_clean],
        ["Hoechster Einzelbetrag EUR", round(highest, 2)],
        ["Durchschnittsbetrag EUR", average],
        ["Groesste Kategorie (Summe)", top_category],
        ["Top-Anbieter (Summe)", top_vendor],
        ["Erstellt am", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]


def _write_csv(path: Path, rows: list[list[object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        for row in rows:
            writer.writerow([_csv_value(v) for v in row])


def _csv_value(value):
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float):
        return f"{value:.2f}".replace(".", ",")
    return value


def _write_xlsx_styled(path: Path, sheets: dict[str, list[list[object]]], year: int) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.formatting.rule import FormulaRule
    except ImportError:
        logging.info("openpyxl nicht installiert, nutze einfachen XLSX-Writer.")
        return False

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = f"Einkommensteuer-Auswertung {year}"
    wb.properties.subject = "Belegauswertung fuer Steuerberater/Finanzamt"
    wb.properties.creator = "Invoice-Agent"

    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="FF305496")
    title_font = Font(bold=True, size=14, color="FF305496")
    review_fill = PatternFill("solid", fgColor="FFFCE4D6")
    missing_fill = PatternFill("solid", fgColor="FFFFF2CC")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="FFE7E6E6")
    label_font = Font(bold=True)
    hyperlink_font = Font(color="FF0563C1", underline="single")
    thin = Side(border_style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    money_fmt = '#,##0.00\u00a0"\u20ac"'
    percent_fmt = '0.00"\u00a0%"'
    int_fmt = "#,##0"
    date_fmt = "DD.MM.YYYY"

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            continue

        # Titel-Zeile oberhalb der Tabelle
        title_text = f"{sheet_name} - Steuerjahr {year}"
        ws.cell(row=1, column=1, value=title_text).font = title_font
        ws.row_dimensions[1].height = 22

        header_row_index = 2
        headers = rows[0]
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row_index, column=col_index, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[header_row_index].height = 20

        amount_columns = _find_amount_columns(headers)
        percent_columns = _find_percent_columns(headers)
        count_columns = _find_count_columns(headers)
        date_columns = _find_date_columns(headers)
        path_columns = _find_path_columns(headers)
        seq_columns = _find_seq_columns(headers)
        status_col = _find_column(headers, "status")
        category_col = _find_column(headers, "steuerkategorie")
        amount_col_for_format = next(iter(amount_columns), None)

        data_start = header_row_index + 1
        for offset, row in enumerate(rows[1:], start=0):
            r = data_start + offset
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=col_index, value=_xlsx_value(value))
                cell.border = border
                if col_index in amount_columns and isinstance(value, (int, float)) and value != "":
                    cell.number_format = money_fmt
                    cell.alignment = right_align
                elif col_index in percent_columns and isinstance(value, (int, float)) and value != "":
                    cell.number_format = percent_fmt
                    cell.alignment = right_align
                elif col_index in count_columns and isinstance(value, (int, float)) and value != "":
                    cell.number_format = int_fmt
                    cell.alignment = right_align
                elif col_index in seq_columns and isinstance(value, (int, float)) and value != "":
                    cell.number_format = int_fmt
                    cell.alignment = center
                elif col_index in date_columns and isinstance(value, date):
                    cell.number_format = date_fmt
                    cell.alignment = center
                else:
                    cell.alignment = left

                # Hyperlink auf Archivpfad / Quelle, wenn Datei existiert
                if col_index in path_columns and isinstance(value, str) and value:
                    try:
                        if Path(value).exists():
                            cell.hyperlink = Path(value).as_uri()
                            cell.font = hyperlink_font
                    except (OSError, ValueError):
                        pass

        # Spaltenbreite
        max_col = len(headers)
        max_row = header_row_index + len(rows) - 1
        for col_index in range(1, max_col + 1):
            letter = get_column_letter(col_index)
            widths = [len(str(_xlsx_value(r[col_index - 1]))) for r in rows if col_index - 1 < len(r)]
            width = max(widths) if widths else 10
            ws.column_dimensions[letter].width = min(max(width + 2, 12), 60)

        # Filter + Freeze (gilt nicht fuer Uebersicht)
        if sheet_name != "Uebersicht":
            ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate
            ws.auto_filter.ref = f"A{header_row_index}:{get_column_letter(max_col)}{max_row}"

        # Bedingte Formatierungen auf "Alle Belege" und "Review"
        if sheet_name in ("Alle Belege", "Review") and category_col and max_row >= data_start:
            cat_letter = get_column_letter(category_col)
            row_range = f"A{data_start}:{get_column_letter(max_col)}{max_row}"
            ws.conditional_formatting.add(
                row_range,
                FormulaRule(formula=[f'EXACT(${cat_letter}{data_start},"Review")'], fill=review_fill),
            )
            if amount_col_for_format:
                amt_letter = get_column_letter(amount_col_for_format)
                ws.conditional_formatting.add(
                    row_range,
                    FormulaRule(formula=[f'ISBLANK(${amt_letter}{data_start})'], fill=missing_fill),
                )

        # Summenzeile mit SUBTOTAL (rechnet mit Filtern korrekt)
        if sheet_name in ("Alle Belege", "Monate", "Quartale", "Kategorien", "Anbieter", "Review") and max_row > header_row_index:
            total_row = max_row + 1
            label_cell = ws.cell(row=total_row, column=1, value="Summe (gefiltert)")
            label_cell.font = total_font
            label_cell.fill = total_fill
            label_cell.border = border
            label_cell.alignment = left
            for col_index in range(2, max_col + 1):
                cell = ws.cell(row=total_row, column=col_index)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
                if col_index in amount_columns:
                    letter = get_column_letter(col_index)
                    cell.value = f"=SUBTOTAL(9,{letter}{data_start}:{letter}{max_row})"
                    cell.number_format = money_fmt
                    cell.alignment = right_align
                elif col_index in count_columns:
                    letter = get_column_letter(col_index)
                    cell.value = f"=SUBTOTAL(9,{letter}{data_start}:{letter}{max_row})"
                    cell.number_format = int_fmt
                    cell.alignment = right_align

        # Druck-Layout
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.7, bottom=0.7)
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = f"{header_row_index}:{header_row_index}"
        ws.oddHeader.left.text = title_text
        ws.oddHeader.right.text = f"Stand: {datetime.now().strftime('%d.%m.%Y')}"
        ws.oddFooter.center.text = "Seite &P von &N"

        # Uebersicht hat besondere Formatierung: linke Spalte fett, ohne Filter
        if sheet_name == "Uebersicht":
            for r in range(data_start, max_row + 1):
                left_cell = ws.cell(row=r, column=1)
                left_cell.font = label_font
                left_cell.alignment = left
                value_cell = ws.cell(row=r, column=2)
                if isinstance(value_cell.value, (int, float)):
                    if "EUR" in str(ws.cell(row=r, column=1).value):
                        value_cell.number_format = money_fmt
                        value_cell.alignment = right_align
                    else:
                        value_cell.number_format = int_fmt
                        value_cell.alignment = right_align
            ws.column_dimensions["A"].width = 38
            ws.column_dimensions["B"].width = 28

    # Uebersicht als erste Registerkarte anzeigen
    if "Uebersicht" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Uebersicht")

    try:
        wb.save(path)
    except Exception as exc:
        logging.warning("openpyxl-Export fehlgeschlagen, nutze Fallback: %s", exc)
        return False
    return True


def _xlsx_value(value):
    if value == "":
        return None
    return value


def _find_column(headers, name: str):
    for i, h in enumerate(headers):
        if isinstance(h, str) and h.lower() == name.lower():
            return i + 1
    return None


def _find_amount_columns(headers) -> set:
    return {
        i + 1
        for i, h in enumerate(headers)
        if isinstance(h, str) and ("betrag" in h.lower() or "summe" in h.lower())
    }


def _find_percent_columns(headers) -> set:
    return {i + 1 for i, h in enumerate(headers) if isinstance(h, str) and "anteil" in h.lower()}


def _find_count_columns(headers) -> set:
    return {i + 1 for i, h in enumerate(headers) if isinstance(h, str) and h.lower() == "anzahl"}


def _find_date_columns(headers) -> set:
    return {i + 1 for i, h in enumerate(headers) if isinstance(h, str) and h.lower() == "datum"}


def _find_path_columns(headers) -> set:
    return {
        i + 1
        for i, h in enumerate(headers)
        if isinstance(h, str) and ("pfad" in h.lower() or h.lower() == "quelle")
    }


def _find_seq_columns(headers) -> set:
    return {i + 1 for i, h in enumerate(headers) if isinstance(h, str) and h.lower().startswith("lfd")}


def _row_is_review(headers, row) -> bool:
    try:
        idx = [h.lower() if isinstance(h, str) else "" for h in headers].index("steuerkategorie")
    except ValueError:
        return False
    return idx < len(row) and str(row[idx]).lower() == "review"


def _write_xlsx(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", _content_types_xml(len(sheets)))
        zf.writestr("_rels/.rels", _rels_xml())
        zf.writestr("xl/workbook.xml", _workbook_xml(list(sheets)))
        zf.writestr("xl/_rels/workbook.xml.rels", _workbook_rels_xml(len(sheets)))
        for index, rows in enumerate(sheets.values(), start=1):
            zf.writestr(f"xl/worksheets/sheet{index}.xml", _worksheet_xml(rows))


def _worksheet_xml(rows: list[list[object]]) -> str:
    sheet_data = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for col_index, value in enumerate(row, start=1):
            ref = f"{_column_name(col_index)}{row_index}"
            if isinstance(value, (int, float)) and value != "":
                cells.append(f'<c r="{ref}"><v>{value}</v></c>')
            else:
                cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>')
        sheet_data.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_data)}</sheetData>'
        "</worksheet>"
    )


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _content_types_xml(sheet_count: int) -> str:
    overrides = [
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
    ]
    for index in range(1, sheet_count + 1):
        overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        f'{"".join(overrides)}'
        "</Types>"
    )


def _rels_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )


def _workbook_xml(sheet_names: list[str]) -> str:
    sheets = []
    for index, name in enumerate(sheet_names, start=1):
        sheets.append(f'<sheet name="{escape(name)}" sheetId="{index}" r:id="rId{index}"/>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets>{"".join(sheets)}</sheets>'
        "</workbook>"
    )


def _workbook_rels_xml(sheet_count: int) -> str:
    relationships = []
    for index in range(1, sheet_count + 1):
        relationships.append(
            f'<Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{index}.xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{"".join(relationships)}'
        "</Relationships>"
    )
