import csv
import logging
import sqlite3
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from xml.sax.saxutils import escape


DEFAULT_CATEGORY_RULES = {
    "all inkl": "Webhosting",
    "all-inkl": "Webhosting",
    "apotheke": "Gesundheit",
    "arzt": "Gesundheit",
    "artz": "Gesundheit",
    "auto": "KFZ",
    "autoversicherung": "Versicherung",
    "congstar": "Telekommunikation",
    "elster": "Steuer",
    "finanzamt": "Steuer",
    "gebäudeversicherung": "Versicherung",
    "gebaeudeversicherung": "Versicherung",
    "grundbesitz": "Immobilie",
    "hausrat": "Versicherung",
    "hotel": "Reisekosten",
    "huk24": "Versicherung",
    "kfz": "KFZ",
    "lkh": "Versicherung",
    "restaurant": "Bewirtung",
    "rechtschutz": "Versicherung",
    "rechtsschutz": "Versicherung",
    "sc technology": "Beratung/Abrechnung",
    "strato": "Webhosting",
    "strom": "Versorgung",
    "tankbeleg": "KFZ",
    "tankbelege": "KFZ",
    "vodafone": "Telekommunikation",
    "wasser": "Versorgung",
    "zahn": "Gesundheit",
    "zahnarzt": "Gesundheit",
}

REVIEW_KEYWORDS = (
    "angebot",
    "anordnung",
    "betriebsprüfung",
    "bescheinigung",
    "elster",
    "login",
    "meldebescheinigung",
    "lohnsteuerbescheinigung",
    "uvg",
)

DETAIL_HEADERS = (
    "Datum",
    "Anbieter",
    "Betrag",
    "Waehrung",
    "Steuerkategorie",
    "Hinweis",
    "Status",
    "Rechnungsnummer",
    "Archivpfad",
    "Quelle",
)

SUMMARY_HEADERS = ("Steuerkategorie", "Anzahl", "Summe EUR")
REVIEW_HEADERS = DETAIL_HEADERS


@dataclass
class TaxExportConfig:
    database_path: Path
    output_dir: Path
    category_rules: dict[str, str]


@dataclass
class TaxExportResult:
    year: int
    rows: int
    review_rows: int
    csv_path: Path
    xlsx_path: Path


def export_tax_year(config: TaxExportConfig, year: int) -> TaxExportResult:
    rows = _load_invoice_rows(config.database_path, year)
    categorized = [_categorize(row, config.category_rules) for row in rows]

    year_dir = config.output_dir / str(year)
    year_dir.mkdir(parents=True, exist_ok=True)
    csv_path = year_dir / f"einkommensteuer_{year}.csv"
    xlsx_path = year_dir / f"einkommensteuer_{year}.xlsx"

    detail_rows = [DETAIL_HEADERS] + [_detail_row(row) for row in categorized]
    summary_rows = [SUMMARY_HEADERS] + _summary_rows(categorized)
    review_rows = [REVIEW_HEADERS] + [_detail_row(row) for row in categorized if row["tax_category"] == "Review"]
    monthly_rows = _monthly_rows(categorized)
    overview_rows = _overview_rows(year, categorized)

    _write_csv(csv_path, detail_rows)

    sheets = {
        "Uebersicht": overview_rows,
        "Alle Belege": detail_rows,
        "Monate": monthly_rows,
        "Kategorien": summary_rows,
        "Review": review_rows,
    }

    if not _write_xlsx_styled(xlsx_path, sheets, year):
        _write_xlsx(xlsx_path, sheets)

    return TaxExportResult(
        year=year,
        rows=len(categorized),
        review_rows=len(review_rows) - 1,
        csv_path=csv_path,
        xlsx_path=xlsx_path,
    )


def _load_invoice_rows(database_path: Path, year: int) -> list[sqlite3.Row]:
    con = sqlite3.connect(database_path)
    con.row_factory = sqlite3.Row
    try:
        return con.execute(
            """
            select invoice_date, vendor, amount, currency, invoice_number, status,
                   source_path, archive_path, category, reason
            from invoices
            where substr(invoice_date, 1, 4) = ?
            order by invoice_date, vendor
            """,
            (str(year),),
        ).fetchall()
    finally:
        con.close()


def _categorize(row: sqlite3.Row, rules: dict[str, str]) -> dict:
    text = " ".join(
        str(row[key] or "")
        for key in ("vendor", "source_path", "archive_path", "category", "reason")
    ).lower()

    tax_category = ""
    notes: list[str] = []

    # Spezifischstes (längstes) Keyword gewinnt, nicht das erste in Dict-Reihenfolge.
    matched = [
        (keyword, category)
        for keyword, category in rules.items()
        if re.search(rf"(?<![a-z0-9]){re.escape(keyword.lower())}(?![a-z0-9])", text)
    ]
    if matched:
        keyword, category = max(matched, key=lambda kc: len(kc[0]))
        tax_category = category
        notes.append(f"Regel: {keyword}")
    else:
        tax_category = "Review"
        notes.append("Keine Regel gefunden")

    if row["status"] != "archived":
        tax_category = "Review"
        notes.append(f"Status: {row['status']}")

    if row["amount"] is None:
        tax_category = "Review"
        notes.append("Betrag fehlt")

    review_matches = [kw for kw in REVIEW_KEYWORDS if kw in text]
    if review_matches:
        tax_category = "Review"
        notes.append("Review-Schluesselwort: " + ", ".join(review_matches[:3]))

    return {
        "invoice_date": row["invoice_date"],
        "vendor": row["vendor"],
        "amount": row["amount"],
        "currency": row["currency"],
        "invoice_number": row["invoice_number"] or "",
        "status": row["status"],
        "source_path": row["source_path"],
        "archive_path": row["archive_path"],
        "tax_category": tax_category,
        "note": "; ".join(notes),
    }


def _detail_row(row: dict) -> list[object]:
    invoice_date = _to_date(row["invoice_date"])
    amount = row["amount"] if row["amount"] is not None else ""
    return [
        invoice_date,
        row["vendor"],
        amount,
        row["currency"],
        row["tax_category"],
        row["note"],
        row["status"],
        row["invoice_number"],
        row["archive_path"],
        row["source_path"],
    ]


def _to_date(value):
    if isinstance(value, date):
        return value
    if not value:
        return ""
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return str(value)


def _summary_rows(rows: Iterable[dict]) -> list[list[object]]:
    grouped: dict[str, dict[str, float]] = {}
    for row in rows:
        if row["currency"] != "EUR" or row["amount"] is None:
            continue
        bucket = grouped.setdefault(row["tax_category"], {"count": 0, "sum": 0.0})
        bucket["count"] += 1
        bucket["sum"] += float(row["amount"])

    output = []
    for category in sorted(grouped):
        bucket = grouped[category]
        output.append([category, bucket["count"], round(bucket["sum"], 2)])
    return output


MONTHLY_HEADERS = ("Monat", "Steuerkategorie", "Anzahl", "Summe EUR")


def _monthly_rows(rows: Iterable[dict]) -> list[list[object]]:
    grouped: dict[tuple[str, str], dict[str, float]] = {}
    for row in rows:
        if row["currency"] != "EUR" or row["amount"] is None:
            continue
        d = _to_date(row["invoice_date"])
        if not isinstance(d, date):
            continue
        month = f"{d.year:04d}-{d.month:02d}"
        key = (month, row["tax_category"])
        bucket = grouped.setdefault(key, {"count": 0, "sum": 0.0})
        bucket["count"] += 1
        bucket["sum"] += float(row["amount"])

    output: list[list[object]] = [list(MONTHLY_HEADERS)]
    for (month, category) in sorted(grouped):
        bucket = grouped[(month, category)]
        output.append([month, category, bucket["count"], round(bucket["sum"], 2)])
    return output


OVERVIEW_HEADERS = ("Kennzahl", "Wert")


def _overview_rows(year: int, rows: list[dict]) -> list[list[object]]:
    total_count = len(rows)
    archived = sum(1 for r in rows if r["status"] == "archived")
    review = sum(1 for r in rows if r["tax_category"] == "Review")
    eur_sum = round(
        sum(float(r["amount"]) for r in rows if r["currency"] == "EUR" and r["amount"] is not None),
        2,
    )
    eur_sum_clean = round(
        sum(
            float(r["amount"])
            for r in rows
            if r["currency"] == "EUR" and r["amount"] is not None and r["tax_category"] != "Review"
        ),
        2,
    )
    missing_amount = sum(1 for r in rows if r["amount"] is None)

    return [
        list(OVERVIEW_HEADERS),
        ["Steuerjahr", year],
        ["Belege gesamt", total_count],
        ["Davon archiviert", archived],
        ["Davon zur Pruefung (Review)", review],
        ["Belege ohne Betrag", missing_amount],
        ["Summe EUR (alle Belege)", eur_sum],
        ["Summe EUR (ohne Review)", eur_sum_clean],
        ["Erstellt am", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]


def _write_csv(path: Path, rows: list[list[object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        for row in rows:
            writer.writerow([_csv_value(v) for v in row])


def _csv_value(value):
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float):
        return f"{value:.2f}".replace(".", ",")
    return value


def _write_xlsx_styled(path: Path, sheets: dict[str, list[list[object]]], year: int) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logging.info("openpyxl nicht installiert, nutze einfachen XLSX-Writer.")
        return False

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF305496")
    review_fill = PatternFill("solid", fgColor="FFFCE4D6")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="FFE7E6E6")
    thin = Side(border_style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    money_fmt = '#,##0.00\u00a0"\u20ac"'
    int_fmt = "#,##0"
    date_fmt = "DD.MM.YYYY"

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            continue

        headers = rows[0]
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_index, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        amount_columns = _find_amount_columns(headers)
        count_columns = _find_count_columns(headers)
        date_columns = _find_date_columns(headers)

        for row_index, row in enumerate(rows[1:], start=2):
            is_review_row = sheet_name == "Alle Belege" and _row_is_review(headers, row)
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=_xlsx_value(value))
                cell.border = border
                cell.alignment = left
                if col_index in amount_columns and isinstance(value, (int, float)) and value != "":
                    cell.number_format = money_fmt
                elif col_index in count_columns and isinstance(value, (int, float)) and value != "":
                    cell.number_format = int_fmt
                elif col_index in date_columns and isinstance(value, date):
                    cell.number_format = date_fmt
                if is_review_row:
                    cell.fill = review_fill

        # Spaltenbreite + Filter + Freeze
        max_col = len(headers)
        max_row = len(rows)
        for col_index in range(1, max_col + 1):
            letter = get_column_letter(col_index)
            width = max(
                (len(str(_xlsx_value(r[col_index - 1]))) for r in rows if col_index - 1 < len(r)),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max(width + 2, 12), 60)
        ws.freeze_panes = "A2"
        if max_row >= 1 and max_col >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # Summenzeile fuer Tabellen mit numerischen Spalten
        if sheet_name in ("Alle Belege", "Monate", "Kategorien") and max_row > 1:
            total_row = max_row + 1
            ws.cell(row=total_row, column=1, value="Summe").font = total_font
            ws.cell(row=total_row, column=1).fill = total_fill
            for col_index in range(1, max_col + 1):
                cell = ws.cell(row=total_row, column=col_index)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
                if col_index in amount_columns:
                    letter = get_column_letter(col_index)
                    cell.value = f"=SUM({letter}2:{letter}{max_row})"
                    cell.number_format = money_fmt
                elif col_index in count_columns:
                    letter = get_column_letter(col_index)
                    cell.value = f"=SUM({letter}2:{letter}{max_row})"
                    cell.number_format = int_fmt

    # Uebersicht als erste Registerkarte anzeigen
    if "Uebersicht" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Uebersicht")

    try:
        wb.save(path)
    except Exception as exc:
        logging.warning("openpyxl-Export fehlgeschlagen, nutze Fallback: %s", exc)
        return False
    return True


def _xlsx_value(value):
    if value == "":
        return None
    return value


def _find_amount_columns(headers) -> set:
    return {
        i + 1
        for i, h in enumerate(headers)
        if isinstance(h, str) and ("betrag" in h.lower() or "summe" in h.lower())
    }


def _find_count_columns(headers) -> set:
    return {i + 1 for i, h in enumerate(headers) if isinstance(h, str) and h.lower() == "anzahl"}


def _find_date_columns(headers) -> set:
    return {i + 1 for i, h in enumerate(headers) if isinstance(h, str) and h.lower() == "datum"}


def _row_is_review(headers, row) -> bool:
    try:
        idx = [h.lower() if isinstance(h, str) else "" for h in headers].index("steuerkategorie")
    except ValueError:
        return False
    return idx < len(row) and str(row[idx]).lower() == "review"


def _write_xlsx(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", _content_types_xml(len(sheets)))
        zf.writestr("_rels/.rels", _rels_xml())
        zf.writestr("xl/workbook.xml", _workbook_xml(list(sheets)))
        zf.writestr("xl/_rels/workbook.xml.rels", _workbook_rels_xml(len(sheets)))
        for index, rows in enumerate(sheets.values(), start=1):
            zf.writestr(f"xl/worksheets/sheet{index}.xml", _worksheet_xml(rows))


def _worksheet_xml(rows: list[list[object]]) -> str:
    sheet_data = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for col_index, value in enumerate(row, start=1):
            ref = f"{_column_name(col_index)}{row_index}"
            if isinstance(value, (int, float)) and value != "":
                cells.append(f'<c r="{ref}"><v>{value}</v></c>')
            else:
                cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>')
        sheet_data.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_data)}</sheetData>'
        "</worksheet>"
    )


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _content_types_xml(sheet_count: int) -> str:
    overrides = [
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
    ]
    for index in range(1, sheet_count + 1):
        overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        f'{"".join(overrides)}'
        "</Types>"
    )


def _rels_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )


def _workbook_xml(sheet_names: list[str]) -> str:
    sheets = []
    for index, name in enumerate(sheet_names, start=1):
        sheets.append(f'<sheet name="{escape(name)}" sheetId="{index}" r:id="rId{index}"/>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets>{"".join(sheets)}</sheets>'
        "</workbook>"
    )


def _workbook_rels_xml(sheet_count: int) -> str:
    relationships = []
    for index in range(1, sheet_count + 1):
        relationships.append(
            f'<Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{index}.xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{"".join(relationships)}'
        "</Relationships>"
    )
